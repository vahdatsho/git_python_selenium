import openpyxl

book = openpyxl.load_workbook("/Users/vahdat/Documents/test_data_Normalization.xlsx")

sheet = book.active

dictionary ={}

cell = sheet.cell(row=2, column=2)
print(cell.value)  # read value

#sheet.cell(row=2, column=2).value = "Vahdat" #write value

print(sheet.cell(row=2, column=2).value)

print(sheet.max_column)
print(sheet.max_row)

#print(sheet['A5'].value) # get by position


print("--------------------------------------------")


for i in range(1, sheet.max_row+1):
    if sheet.cell(row=i, column=1).value == "testcase 2":
        for j in range(2, sheet.max_column+1):

            dictionary[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value


"""
for i in range(1, sheet.max_row):
    if sheet.cell(row=i, column=1).value == "testcase 2":
        for j in range(1, sheet.max_column+1):
            print(sheet.cell(row=i, column=j).value)
"""