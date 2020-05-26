import openpyxl


class HomePageData:

        test_HomePage_data = [{"firstname": "Amir", "email": "vahdat@hulu.com", "gender": "Male"}, {"firstname": "Munira", "email": "munir@hulu.com", "gender": "Female"}]


        @staticmethod
        def get_test_data(test_case_name):

                dictionary = {}

                book = openpyxl.load_workbook("/Users/vahdat/Documents/test_data_Normalization.xlsx")

                sheet = book.active

                for i in range(1, sheet.max_row + 1):
                        if sheet.cell(row=i, column=1).value == test_case_name:
                                for j in range(2, sheet.max_column + 1):
                                        dictionary[sheet.cell(row=1, column=j).value] = sheet.cell(row=i,
                                                                                                   column=j).value

                return [dictionary]
